from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.pagebreak import Break
import math
from openpyxl.worksheet.page import PageMargins
from flask import Blueprint, render_template, request, redirect, jsonify, flash, send_file, session
from app.routes.main_routes import login_required
from collections import defaultdict
from app.config import get_db_connection
import unicodedata

viagens = Blueprint("viagens", __name__)
def parse_data(data_str):
    if not data_str:
        return None
    try:
        return datetime.strptime(data_str, "%Y-%m-%d").date()
    except ValueError:
        return None
    
# listar
@viagens.route("/viagens")
@login_required
def listar_viagens():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    data = parse_data(request.args.get("data"))
    data_inicio = parse_data(request.args.get("data_inicio"))
    data_fim = parse_data(request.args.get("data_fim"))
    local = request.args.get("local")

    query = """
        SELECT *
        FROM viagens
        WHERE ativo = 1
    """

    params = []

    if local:
        query += " AND local = %s"
        params.append(local)

    if data:
        query += " AND data_viagem = %s"
        params.append(data)
        

    elif data_inicio and data_fim:
        query += " AND data_viagem BETWEEN %s AND %s"
        params.append(data_inicio)
        params.append(data_fim)

    elif data_inicio:
        query += " AND data_viagem >= %s"
        params.append(data_inicio)

    elif data_fim:
        query += " AND data_viagem <= %s"
        params.append(data_fim)

    query += """
    ORDER BY 
        CASE 
            WHEN status = 'Em andamento' THEN 1
            WHEN status = 'Planejada' THEN 2
            WHEN status = 'Finalizada' THEN 3
            ELSE 4
        END,
        data_viagem DESC
    """

    cursor.execute(query, tuple(params))
    viagens_lista = cursor.fetchall()


    cursor.close()
    conn.close()

    filtrado = bool(data or data_inicio or data_fim or local)

    data_invalida = (
    (request.args.get("data") and not data) or
    (request.args.get("data_inicio") and not data_inicio) or
    (request.args.get("data_fim") and not data_fim)
    )

    return render_template(
        "viagens.html",
        viagens=viagens_lista,
        filtrado=filtrado,
        data_invalida=data_invalida
    )


# nova viagem
@viagens.route("/viagens/nova", methods=["GET","POST"])
@login_required
def nova_viagem():

    if request.method == "POST":

        local = request.form["local"]
        data_str = request.form.get("data_viagem")
        try:
            data_viagem = datetime.strptime(data_str, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            flash("Data inválida!")
            return redirect(request.url)
        observacoes = request.form["observacoes"].strip()

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO viagens
            (local, data_viagem, observacoes, status)
            VALUES (%s,%s,%s,%s)
        """, (local, data_viagem, observacoes, "Planejada"))

        conn.commit()

        viagem_id = cursor.lastrowid

        cursor.close()
        conn.close()

        flash("Viagem criada com sucesso!", "success")

        return redirect(f"/viagens/{viagem_id}")

    return render_template("nova_viagem.html")

# detalhes da viagem
@viagens.route("/viagens/<int:id>")
@login_required
def detalhe_viagem(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)


    cursor.execute("""
        SELECT *
        FROM viagens
        WHERE id = %s
    """, (id,))

    viagem = cursor.fetchone()


    cursor.execute("""
        SELECT 
            pedidos.id,
            pedidos.ordem,
            pedidos.tipo,
            pedidos.observacoes,
            clientes.nome AS cliente_nome,
            clientes.ativo AS cliente_ativo,
            lojas.id AS loja_id,
            lojas.nome AS loja_nome,
            lojas.ativo AS loja_ativo,
            shopping.id AS shopping_id,
            shopping.nome AS shopping_nome,
            shopping.ativo AS shopping_ativo,
            vl.id AS viagem_loja_id,
            vs.id AS viagem_shopping_id

        FROM pedidos

        LEFT JOIN clientes 
        ON pedidos.cliente_id = clientes.id

        LEFT JOIN lojas 
        ON pedidos.loja_id = lojas.id

        LEFT JOIN shopping 
        ON lojas.shopping_id = shopping.id

        LEFT JOIN viagem_loja vl 
        ON vl.loja_id = lojas.id 
        AND vl.viagem_id = pedidos.viagem_id   

        LEFT JOIN viagem_shopping vs 
        ON vs.shopping_id = shopping.id 
        AND vs.viagem_id = pedidos.viagem_id        

        WHERE pedidos.viagem_id = %s

        ORDER BY 
        COALESCE(vs.ordem, 9999),
        COALESCE(vl.ordem, 9999),
        pedidos.ordem
    """, (id,))

    pedidos = cursor.fetchall()

    estrutura = {}

    for pedido in pedidos:

        shopping = pedido["shopping_nome"] or "Sem shopping"
        loja = pedido["loja_nome"] or "Sem loja"

        if shopping not in estrutura:
            estrutura[shopping] = {
                "id": pedido["viagem_shopping_id"],
                "shopping_id": pedido["shopping_id"],
                "lojas": {}
            }

        if loja not in estrutura[shopping]["lojas"]:
            estrutura[shopping]["lojas"][loja] = {
                "id": pedido["viagem_loja_id"],
                "loja_id": pedido["loja_id"],
                "pedidos": []
            }

        estrutura[shopping]["lojas"][loja]["pedidos"].append(pedido)

    cursor.execute("""
    SELECT vc.id, vc.ordem, c.nome, c.endereco, c.ativo AS cliente_ativo, vc.cliente_id
    FROM viagem_clientes vc
    LEFT JOIN clientes c ON c.id = vc.cliente_id
    WHERE vc.viagem_id = %s
    ORDER BY vc.ordem
""", (id,))

    clientes_viagem = cursor.fetchall()

    cursor.execute("""
    SELECT id, nome 
    FROM clientes 
    WHERE ativo = 1
    ORDER BY LOWER(nome) ASC
    """)

    clientes = cursor.fetchall()

    cursor.execute("""
    SELECT *
    FROM viagem_financeiro
    WHERE viagem_id = %s
""", (id,))

    financeiro = cursor.fetchall()

    total_ganho = sum(f["valor"] for f in financeiro if f["tipo"] == "GANHO")
    total_custo = sum(f["valor"] for f in financeiro if f["tipo"] == "CUSTO")

    saldo = total_ganho - total_custo

    total_clientes = len(clientes_viagem)

    total_shoppings = len(estrutura)

    total_lojas = sum(len(shopping["lojas"]) for shopping in estrutura.values())

    total_pedidos = sum(
    len(dados_loja["pedidos"])
    for dados in estrutura.values()
    for dados_loja in dados["lojas"].values()
    )

    cursor.close()
    conn.close()

    return render_template(
        "viagem_detalhe.html",
        viagem=viagem,
        estrutura=estrutura,
        clientes_viagem=clientes_viagem,
        clientes=clientes,
        financeiro=financeiro,
        total_ganho=total_ganho,
        total_custo=total_custo,
        saldo=saldo,
        total_clientes=total_clientes,
        total_shoppings=total_shoppings,
        total_lojas=total_lojas,
        total_pedidos=total_pedidos
    )

#botao subir shop
@viagens.route("/shopping/<int:id>/subir", methods=["GET"])
@login_required
def subir_shopping(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT * FROM viagem_shopping
        WHERE id = %s
    """, (id,))
    atual = cursor.fetchone()

    cursor.execute("""
        SELECT * FROM viagem_shopping
        WHERE viagem_id = %s
        AND ordem < %s
        ORDER BY ordem DESC
        LIMIT 1
    """, (atual["viagem_id"], atual["ordem"]))

    anterior = cursor.fetchone()

    if anterior:
        cursor.execute(
            "UPDATE viagem_shopping SET ordem=%s WHERE id=%s",
            (anterior["ordem"], atual["id"])
        )

        cursor.execute(
            "UPDATE viagem_shopping SET ordem=%s WHERE id=%s",
            (atual["ordem"], anterior["id"])
        )

        conn.commit()

    cursor.close()
    conn.close()

    return redirect(f"/viagens/{atual['viagem_id']}")

#descer shop
@viagens.route("/shopping/<int:id>/descer", methods=["GET"])
@login_required
def descer_shopping(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT * FROM viagem_shopping
        WHERE id = %s
    """, (id,))
    atual = cursor.fetchone()

    cursor.execute("""
        SELECT * FROM viagem_shopping
        WHERE viagem_id = %s
        AND ordem > %s
        ORDER BY ordem ASC
        LIMIT 1
    """, (atual["viagem_id"], atual["ordem"]))

    proximo = cursor.fetchone()

    if proximo:
        cursor.execute(
            "UPDATE viagem_shopping SET ordem=%s WHERE id=%s",
            (proximo["ordem"], atual["id"])
        )

        cursor.execute(
            "UPDATE viagem_shopping SET ordem=%s WHERE id=%s",
            (atual["ordem"], proximo["id"])
        )

        conn.commit()

    cursor.close()
    conn.close()

    return redirect(f"/viagens/{atual['viagem_id']}")

#subir botao loja
@viagens.route("/loja/<int:id>/subir", methods=["GET"])
@login_required
def subir_loja(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT * FROM viagem_loja
        WHERE id = %s
    """, (id,))
    atual = cursor.fetchone()

    cursor.execute("""
        SELECT * FROM viagem_loja
        WHERE viagem_id = %s
        AND shopping_id = %s
        AND ordem < %s
        ORDER BY ordem DESC
        LIMIT 1
    """, (atual["viagem_id"], atual["shopping_id"], atual["ordem"]))

    anterior = cursor.fetchone()

    if anterior:
        cursor.execute(
            "UPDATE viagem_loja SET ordem=%s WHERE id=%s",
            (anterior["ordem"], atual["id"])
        )

        cursor.execute(
            "UPDATE viagem_loja SET ordem=%s WHERE id=%s",
            (atual["ordem"], anterior["id"])
        )

        conn.commit()

    cursor.close()
    conn.close()

    return redirect(f"/viagens/{atual['viagem_id']}")

#descer loja
@viagens.route("/loja/<int:id>/descer", methods=["GET"])
@login_required
def descer_loja(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT * FROM viagem_loja
        WHERE id = %s
    """, (id,))
    atual = cursor.fetchone()

    cursor.execute("""
        SELECT * FROM viagem_loja
        WHERE viagem_id = %s
        AND shopping_id = %s
        AND ordem > %s
        ORDER BY ordem ASC
        LIMIT 1
    """, (atual["viagem_id"], atual["shopping_id"], atual["ordem"]))

    proximo = cursor.fetchone()

    if proximo:
        cursor.execute(
            "UPDATE viagem_loja SET ordem=%s WHERE id=%s",
            (proximo["ordem"], atual["id"])
        )

        cursor.execute(
            "UPDATE viagem_loja SET ordem=%s WHERE id=%s",
            (atual["ordem"], proximo["id"])
        )

        conn.commit()

    cursor.close()
    conn.close()

    return redirect(f"/viagens/{atual['viagem_id']}")

#registrar pedido na viagem
@viagens.route("/viagens/<int:id>/novo_pedido", methods=["GET","POST"])
@login_required
def novo_pedido(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":

        loja_id = request.form["loja_id"]
        cliente_id = request.form["cliente_id"]
        tipos = request.form.getlist("tipo[]")
        if not tipos:
            flash("Selecione pelo menos um tipo.", "error")
            return redirect(request.url)

        tipo = ",".join(tipos)
        observacoes = request.form.get("observacoes","").strip()

        cursor.execute("""
            SELECT shopping_id FROM lojas WHERE id = %s
        """, (loja_id,))

        shopping_id = cursor.fetchone()["shopping_id"]

        session["ultimo_loja_id"] = int(loja_id)
        session["ultimo_cliente_id"] = int(cliente_id)
        session["ultimo_shopping_id"] = int(shopping_id)

        cursor.execute("""
            SELECT id FROM viagem_shopping
            WHERE viagem_id = %s AND shopping_id = %s
        """, (id, shopping_id))

        vs = cursor.fetchone()

        if not vs:
            cursor.execute("""
                SELECT COALESCE(MAX(ordem),0)+1 AS ordem
                FROM viagem_shopping
                WHERE viagem_id = %s
            """, (id,))
            ordem_vs = cursor.fetchone()["ordem"]

            cursor.execute("""
                INSERT INTO viagem_shopping (viagem_id, shopping_id, ordem)
                VALUES (%s,%s,%s)
            """, (id, shopping_id, ordem_vs))

        cursor.execute("""
            SELECT id FROM viagem_loja
            WHERE viagem_id = %s AND loja_id = %s
        """, (id, loja_id))

        vl = cursor.fetchone()

        if not vl:
            cursor.execute("""
                SELECT COALESCE(MAX(ordem),0)+1 AS ordem
                FROM viagem_loja
                WHERE viagem_id = %s AND shopping_id = %s
            """, (id, shopping_id))
            ordem_vl = cursor.fetchone()["ordem"]

            cursor.execute("""
                INSERT INTO viagem_loja (viagem_id, shopping_id, loja_id, ordem)
                VALUES (%s,%s,%s,%s)
            """, (id, shopping_id, loja_id, ordem_vl))

        cursor.execute("""
            SELECT COALESCE(MAX(ordem),0)+1 AS nova_ordem
            FROM pedidos
            WHERE viagem_id = %s
            AND loja_id = %s
        """, (id, loja_id))

        ordem = cursor.fetchone()["nova_ordem"]

        cursor.execute("""
            INSERT INTO pedidos
            (viagem_id, loja_id, cliente_id, tipo, observacoes, ordem)
            VALUES (%s,%s,%s,%s,%s,%s)
        """,(id, loja_id, cliente_id, tipo, observacoes, ordem))

        conn.commit()

        cursor.close()
        conn.close()

        flash("Tarefa adicionada com sucesso!", "success")

        return redirect(f"/viagens/{id}")

    cursor.execute("SELECT id,nome FROM shopping WHERE ativo=1 ORDER BY LOWER(nome) ASC")
    shoppings = cursor.fetchall()

    cursor.execute("SELECT id,nome,shopping_id FROM lojas WHERE ativo=1 ORDER BY shopping_id, LOWER(nome) ASC")
    lojas = cursor.fetchall()

    cursor.execute("SELECT id,nome FROM clientes WHERE ativo=1 ORDER BY LOWER(nome) ASC")
    clientes = cursor.fetchall()

    ultimo_loja_id = session.get("ultimo_loja_id")
    ultimo_cliente_id = session.get("ultimo_cliente_id")
    ultimo_shopping_id = session.get("ultimo_shopping_id")

    cursor.close()
    conn.close()

    return render_template(
        "novo_pedido.html",
        viagem_id=id,
        shoppings=shoppings,
        lojas=lojas,
        clientes=clientes,
        ultimo_loja_id=ultimo_loja_id,
        ultimo_cliente_id=ultimo_cliente_id,
        ultimo_shopping_id=ultimo_shopping_id
        )

#status viagem
@viagens.route("/viagens/<int:id>/status", methods=["POST"])
@login_required
def atualizar_status_viagem(id):

    data = request.get_json()
    status = data["status"]

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE viagens
        SET status = %s
        WHERE id = %s
    """, (status, id))

    conn.commit()

    cursor.close()
    conn.close()

    return jsonify({"success": True})

#editar viagem
@viagens.route("/viagens/<int:id>/editar", methods=["GET","POST"])
@login_required
def editar_viagem(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":

        local = request.form["local"]
        data_viagem = request.form["data_viagem"]
        observacoes = request.form["observacoes"].strip()

        cursor.execute("""
            UPDATE viagens
            SET local=%s, data_viagem=%s, observacoes=%s
            WHERE id=%s
        """,(local,data_viagem,observacoes,id))

        conn.commit()

        cursor.close()
        conn.close()

        flash("Viagem atualizada com sucesso!", "success")

        return redirect("/viagens")


    cursor.execute("""
        SELECT *
        FROM viagens
        WHERE id=%s
    """,(id,))

    viagem = cursor.fetchone()

    cursor.close()
    conn.close()

    return render_template("nova_viagem.html", viagem=viagem)

#excluir viagem
@viagens.route("/viagens/<int:id>/excluir")
@login_required
def excluir_viagem(id):

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE viagens
        SET ativo = 0
        WHERE id = %s
    """,(id,))

    conn.commit()

    cursor.close()
    conn.close()

    flash("Viagem excluída com sucesso!", "success")

    return redirect("/viagens")


#seta p cima pedido
@viagens.route("/pedidos/<int:id>/subir", methods=["POST"])
@login_required
def subir_pedido(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id, viagem_id, loja_id, ordem
        FROM pedidos
        WHERE id = %s
    """,(id,))
    pedido = cursor.fetchone()

    cursor.execute("""
        SELECT id, ordem
        FROM pedidos
        WHERE viagem_id = %s 
        AND loja_id = %s
        AND ordem < %s
        ORDER BY ordem DESC
        LIMIT 1
    """,(pedido["viagem_id"], pedido["loja_id"], pedido["ordem"]))

    anterior = cursor.fetchone()

    if anterior:

        cursor.execute(
            "UPDATE pedidos SET ordem=%s WHERE id=%s",
            (anterior["ordem"], pedido["id"])
        )

        cursor.execute(
            "UPDATE pedidos SET ordem=%s WHERE id=%s",
            (pedido["ordem"], anterior["id"])
        )

        conn.commit()

    cursor.close()
    conn.close()

    return ("", 204)

#botao descer seta pedido
@viagens.route("/pedidos/<int:id>/descer", methods=["POST"])
@login_required
def descer_pedido(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id, viagem_id, loja_id, ordem
        FROM pedidos
        WHERE id = %s
    """,(id,))
    pedido = cursor.fetchone()

    cursor.execute("""
        SELECT id, ordem
        FROM pedidos
        WHERE viagem_id = %s
        AND loja_id = %s
        AND ordem > %s
        ORDER BY ordem ASC
        LIMIT 1
    """,(pedido["viagem_id"], pedido["loja_id"], pedido["ordem"]))

    proximo = cursor.fetchone()

    if proximo:


        cursor.execute(
            "UPDATE pedidos SET ordem=%s WHERE id=%s",
            (proximo["ordem"], pedido["id"])
        )

        cursor.execute(
            "UPDATE pedidos SET ordem=%s WHERE id=%s",
            (pedido["ordem"], proximo["id"])
        )

        conn.commit()

    cursor.close()
    conn.close()

    return ("", 204)


#editar pedido
@viagens.route("/pedidos/<int:id>/editar", methods=["GET","POST"])
@login_required
def editar_pedido(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":

        loja_id = request.form["loja_id"]
        cliente_id = request.form["cliente_id"]
        tipos = request.form.getlist("tipo[]")

        if not tipos:
            flash("Selecione pelo menos um tipo.", "error")
            return redirect(request.url)

        tipo = ",".join(tipos)
        observacoes = request.form.get("observacoes","").strip()
        next_url = request.form.get("next")

        cursor.execute("""
            UPDATE pedidos
            SET loja_id=%s, cliente_id=%s, tipo=%s, observacoes=%s
            WHERE id=%s
        """,(loja_id, cliente_id, tipo, observacoes, id))

        conn.commit()

        cursor.execute("SELECT viagem_id FROM pedidos WHERE id=%s",(id,))
        viagem_id = cursor.fetchone()["viagem_id"]

        cursor.close()
        conn.close()

        flash("Tarefa atualizada com sucesso!", "success")

        if next_url:
            return redirect(next_url)

        return redirect(f"/viagens/{viagem_id}")

    cursor.execute("""
        SELECT *
        FROM pedidos
        WHERE id=%s
    """,(id,))
    pedido = cursor.fetchone()

    viagem_id = pedido["viagem_id"]

    cursor.execute("SELECT id,nome FROM shopping WHERE ativo=1 ORDER BY LOWER(nome) ASC")
    shoppings = cursor.fetchall()

    cursor.execute("SELECT id,nome,shopping_id FROM lojas WHERE ativo=1 ORDER BY shopping_id, LOWER(nome) ASC")
    lojas = cursor.fetchall()

    cursor.execute("SELECT id,nome FROM clientes WHERE ativo=1 ORDER BY LOWER(nome) ASC")
    clientes = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        "novo_pedido.html",
        pedido=pedido,
        viagem_id=viagem_id,
        shoppings=shoppings,
        lojas=lojas,
        clientes=clientes
    )

#botao apagar pedido
@viagens.route("/pedidos/<int:id>/excluir")
@login_required
def excluir_pedido(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT viagem_id, loja_id FROM pedidos WHERE id=%s",(id,))
    pedido = cursor.fetchone()

    viagem_id = pedido["viagem_id"]
    loja_id = pedido["loja_id"]

    cursor.execute("DELETE FROM pedidos WHERE id=%s",(id,))

    cursor.execute("SET @ordem = 0")

    cursor.execute("""
        UPDATE pedidos
        SET ordem = (@ordem := @ordem + 1)
        WHERE viagem_id = %s
        AND loja_id = %s
        ORDER BY ordem
    """, (viagem_id, loja_id))

    conn.commit()

    cursor.close()
    conn.close()

    flash("Tarefa excluída com sucesso!", "success")

    return redirect(f"/viagens/{viagem_id}")

#add cliente pra ordem
@viagens.route("/viagens/<int:id>/add_cliente", methods=["POST"])
@login_required
def add_cliente(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cliente_id = request.form["cliente_id"]


    cursor.execute("""
        SELECT 1 FROM viagem_clientes
        WHERE viagem_id=%s AND cliente_id=%s
    """, (id, cliente_id))

    if cursor.fetchone():
        flash("Este cliente já está na viagem!", "error")
        cursor.close()
        conn.close()
        return redirect(request.form.get("next") or f"/viagens/{id}")


    cursor.execute("""
        SELECT COALESCE(MAX(ordem),0)+1 AS nova_ordem
        FROM viagem_clientes
        WHERE viagem_id = %s
    """, (id,))

    ordem = cursor.fetchone()["nova_ordem"]

    cursor.execute("""
        INSERT INTO viagem_clientes (viagem_id, cliente_id, ordem)
        VALUES (%s,%s,%s)
    """, (id, cliente_id, ordem))

    conn.commit()

    cursor.close()
    conn.close()

    flash("Cliente adicionado na ordem!", "success")

    return redirect(request.form.get("next") or f"/viagens/{id}")


#exccluir da lista
@viagens.route("/viagem_cliente/<int:id>/excluir")
@login_required
def excluir_cliente_viagem(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    next_url = request.args.get("next")

    cursor.execute("SELECT viagem_id FROM viagem_clientes WHERE id=%s", (id,))
    data = cursor.fetchone()
    viagem_id = data["viagem_id"]

    cursor.execute("DELETE FROM viagem_clientes WHERE id=%s", (id,))

    cursor.execute("SET @ordem = 0")

    cursor.execute("""
        UPDATE viagem_clientes
        SET ordem = (@ordem := @ordem + 1)
        WHERE viagem_id = %s
        ORDER BY ordem
    """, (viagem_id,))

    conn.commit()

    cursor.close()
    conn.close()

    flash("Cliente removido com sucesso!", "success")

    if next_url:
        return redirect(next_url)

    return redirect(f"/viagens/{viagem_id}")

#seta subir cliente ordem
@viagens.route("/viagem_cliente/<int:id>/subir")
@login_required
def subir_cliente(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM viagem_clientes WHERE id=%s", (id,))
    atual = cursor.fetchone()

    cursor.execute("""
        SELECT * FROM viagem_clientes
        WHERE viagem_id=%s AND ordem < %s
        ORDER BY ordem DESC LIMIT 1
    """, (atual["viagem_id"], atual["ordem"]))

    anterior = cursor.fetchone()

    if anterior:
        cursor.execute("UPDATE viagem_clientes SET ordem=%s WHERE id=%s",
                       (anterior["ordem"], atual["id"]))

        cursor.execute("UPDATE viagem_clientes SET ordem=%s WHERE id=%s",
                       (atual["ordem"], anterior["id"]))

        conn.commit()

    cursor.close()
    conn.close()

    return redirect(request.args.get("next") or f"/viagens/{atual['viagem_id']}")


#descer ordem
@viagens.route("/viagem_cliente/<int:id>/descer")
@login_required
def descer_cliente(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM viagem_clientes WHERE id=%s", (id,))
    atual = cursor.fetchone()

    cursor.execute("""
        SELECT * FROM viagem_clientes
        WHERE viagem_id=%s AND ordem > %s
        ORDER BY ordem ASC LIMIT 1
    """, (atual["viagem_id"], atual["ordem"]))

    proximo = cursor.fetchone()

    if proximo:
        cursor.execute("UPDATE viagem_clientes SET ordem=%s WHERE id=%s",
                       (proximo["ordem"], atual["id"]))

        cursor.execute("UPDATE viagem_clientes SET ordem=%s WHERE id=%s",
                       (atual["ordem"], proximo["id"]))

        conn.commit()

    cursor.close()
    conn.close()

    return redirect(request.args.get("next") or f"/viagens/{atual['viagem_id']}")

#add financeiro
@viagens.route("/viagens/<int:id>/financeiro/add", methods=["POST"])
@login_required
def add_financeiro(id):

    tipo = request.form["tipo"]
    valor = float(request.form["valor"])
    descricao = request.form.get("descricao", "")

    if valor <= 0:
        flash("O valor deve ser positivo!", "error")
        return jsonify({"erro": "Valor inválido"}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        INSERT INTO viagem_financeiro
        (viagem_id, tipo, valor, descricao)
        VALUES (%s,%s,%s,%s)
    """, (id, tipo, valor, descricao))

    conn.commit()

    cursor.execute("SELECT LAST_INSERT_ID() as id")
    novo_id = cursor.fetchone()["id"]

    cursor.close()
    conn.close()


    return jsonify({
    "id": novo_id,
    "tipo": tipo,
    "valor": valor,
    "descricao": descricao
    })

#deletar financeiro
@viagens.route("/financeiro/<int:id>/delete", methods=["POST"])
@login_required
def deletar_financeiro(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # pega antes de deletar (pra atualizar saldo)
    cursor.execute("SELECT tipo, valor FROM viagem_financeiro WHERE id = %s", (id,))
    item = cursor.fetchone()

    if not item:
        return jsonify({"erro": "Não encontrado"}), 404

    cursor.execute("DELETE FROM viagem_financeiro WHERE id = %s", (id,))
    conn.commit()

    cursor.close()
    conn.close()

    return jsonify({
        "tipo": item["tipo"],
        "valor": item["valor"]
    })

#tabelas export

import io
@viagens.route("/viagens/<int:id>/exportar_tarefas")
@login_required
def exportar_tarefas(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM viagens WHERE id=%s", (id,))
    viagem = cursor.fetchone()

    cursor.execute("""
        SELECT
            pedidos.ordem,
            pedidos.tipo,
            clientes.nome AS cliente_nome,
            clientes.cpf_cnpj,
            lojas.nome AS loja_nome,
            shopping.nome AS shopping_nome,
            lojas.id AS loja_id,
            shopping.id AS shopping_id
        FROM pedidos
        LEFT JOIN clientes   ON pedidos.cliente_id  = clientes.id
        LEFT JOIN lojas      ON pedidos.loja_id     = lojas.id
        LEFT JOIN shopping   ON lojas.shopping_id   = shopping.id
        LEFT JOIN viagem_loja vl
            ON vl.loja_id    = lojas.id
            AND vl.viagem_id = pedidos.viagem_id
        LEFT JOIN viagem_shopping vs
            ON vs.shopping_id = shopping.id
            AND vs.viagem_id  = pedidos.viagem_id
        WHERE pedidos.viagem_id = %s
        ORDER BY
            COALESCE(vs.ordem, 9999),
            COALESCE(vl.ordem, 9999),
            pedidos.ordem
    """, (id,))

    pedidos = cursor.fetchall()
    cursor.close()
    conn.close()

    if viagem.get("local") == "Santa Catarina":
        return _exportar_santa_catarina(viagem, pedidos)
    else:
        return _exportar_serra(viagem, pedidos)


# ── UTILITÁRIOS ───────────────────────────────────────────────────────────────

def _limpar_texto(texto):
    import unicodedata
    return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')


def _nome_arquivo(viagem):
    local = viagem.get("local") or "viagem"
    data  = viagem.get("data_viagem")
    local_str = _limpar_texto(local).replace(" ", "_").lower()
    data_str  = data.strftime("%d-%m-%Y") if data else "sem_data"
    return f"Viagem_{local_str}_{data_str}.xlsx"


def _organizar_estrutura(pedidos):
    estrutura       = defaultdict(lambda: defaultdict(list))
    ordem_shoppings = []
    ordem_lojas     = {}

    for p in pedidos:
        shopping_id   = p.get("shopping_id") or 0
        loja_id       = p.get("loja_id") or 0
        shopping_nome = (p["shopping_nome"] or "-").strip() or "-"
        loja_nome     = (p["loja_nome"] or "-").strip() or "-"

        if shopping_id not in estrutura:
            ordem_shoppings.append((shopping_id, shopping_nome))

        if shopping_id not in ordem_lojas:
            ordem_lojas[shopping_id] = []

        if loja_id not in estrutura[shopping_id]:
            ordem_lojas[shopping_id].append((loja_id, loja_nome))

        estrutura[shopping_id][loja_id].append(p)

    return estrutura, ordem_shoppings, ordem_lojas


def _configurar_colunas(ws):
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 24


def _configurar_impressao(ws, ultima_linha):
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered   = False
    ws.page_setup.fitToHeight  = False
    ws.page_setup.fitToWidth   = False
    ws.page_setup.fitToPage    = False
    ws.page_setup.scale        = 100
    ws.page_setup.orientation  = ws.ORIENTATION_PORTRAIT
    ws.print_area = f"A1:D{ultima_linha}"
    ws.page_margins = PageMargins(
    left=0.10, right=0.10, top=0.10, bottom=0.10,
    header=0, footer=0
    )
    ws.oddHeader.center.text = ""
    ws.oddFooter.center.text = ""


def _calcular_altura_linha(textos, larguras, altura_base):
    maior = 1
    for texto, larg in zip(textos, larguras):
        if not texto:
            continue
        linhas = math.ceil(len(str(texto)) / larg)
        if linhas > maior:
            maior = linhas
    return altura_base * max(1, maior)


def _escrever_linha_pedido(ws, row, loja_nome, p, fonte_padrao, altura_base):
    cliente = (p["cliente_nome"] or "-").upper()
    tipo    = (p["tipo"] or "-")
    tipo    = tipo.replace(",", ", ").replace("_", "/").upper() if tipo != "-" else "-"
    doc     = p["cpf_cnpj"] or "-"

    cel_loja    = ws.cell(row=row, column=1, value=(loja_nome or "-").upper())
    cel_cliente = ws.cell(row=row, column=2, value=cliente)
    cel_tipo    = ws.cell(row=row, column=3, value=tipo)
    cel_doc     = ws.cell(row=row, column=4, value=doc)

    for cel, hor in [(cel_loja, "left"), (cel_cliente, "left"),
                     (cel_tipo, "center"), (cel_doc, "center")]:
        cel.alignment = Alignment(wrap_text=True, horizontal=hor, vertical="center")
        cel.font = fonte_padrao

    altura = _calcular_altura_linha(
        [loja_nome, cliente, tipo, doc],
        [24, 32, 20, 22],
        altura_base
    )
    ws.row_dimensions[row].height = altura
    return altura


def _escrever_header(ws, row, viagem, altura_base):
    local    = (viagem.get("local") or "-").upper()
    data_obj = viagem.get("data_viagem")
    data     = data_obj.strftime("%d/%m/%Y") if data_obj else "-"

    fonte_titulo = Font(size=14, bold=True)
    fill_cinza   = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    ws.cell(row=row, column=1, value=local).font = fonte_titulo
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
    ws.cell(row=row, column=3, value=data).font = fonte_titulo
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = fill_cinza

    ws.row_dimensions[row].height = altura_base
    return row + 1


def _aplicar_bordas_pagina(ws, linha_inicio, linha_fim, linhas_sem_borda=None):
    borda = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    linhas_sem_borda = linhas_sem_borda or set()
    for r in range(linha_inicio, linha_fim + 1):
        if r in linhas_sem_borda:
            continue
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = borda


def _preencher_ate_fim_pagina(ws, row, linha_inicio_pagina, linhas_pagina, altura_base):
    """Adiciona linhas vazias até completar a página inteira."""
    linhas_usadas = row - linha_inicio_pagina
    linhas_faltam = linhas_pagina - linhas_usadas
    for _ in range(max(0, linhas_faltam)):
        ws.row_dimensions[row].height = altura_base
        row += 1
    return row


# ── SERRA ─────────────────────────────────────────────────────────────────────

def _exportar_serra(viagem, pedidos):
    ALTURA_BASE   = 18
    LINHAS_PAGINA = 42
    LINHAS_HEADER = 4   # 2 branco + 1 título + 1 espaço
    LINHAS_UTEIS  = LINHAS_PAGINA - LINHAS_HEADER
    ESPACO_SHOP   = 2

    estrutura, ordem_shoppings, ordem_lojas = _organizar_estrutura(pedidos)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tarefas"
    _configurar_colunas(ws)

    fonte_padrao     = Font(size=13)
    fill_shop        = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    linhas_sem_borda = set()
    paginas_bordas   = []

    def _escrever_cabecalho_pagina(row):
        """Escreve 2 linhas branco + header + 1 espaço. Retorna próxima row."""
        for r in (row, row + 1):
            ws.row_dimensions[r].height = ALTURA_BASE
            linhas_sem_borda.add(r)
        row += 2
        row = _escrever_header(ws, row, viagem, ALTURA_BASE)  # escreve 1 linha, row+1
        ws.row_dimensions[row].height = ALTURA_BASE
        row += 1
        return row

    def _tam_shopping(shopping_id):
        """Quantas linhas esse shopping vai ocupar (nome + pedidos + espaço)."""
        total = 1  # linha do nome do shopping
        for loja_id, loja_nome in ordem_lojas[shopping_id]:
            for p in estrutura[shopping_id][loja_id]:
                cliente = (p["cliente_nome"] or "-").upper()
                tipo    = (p["tipo"] or "-").replace(",", ", ").replace("_", "/").upper()
                doc     = p["cpf_cnpj"] or "-"
                h = _calcular_altura_linha(
                    [loja_nome, cliente, tipo, doc],
                    [24, 32, 20, 22], ALTURA_BASE)
                total += max(1, round(h / ALTURA_BASE))
        total += ESPACO_SHOP
        return total

    def _escrever_shopping(row, shopping_id, shopping_nome):
        linhas = 1  # linha do nome
        cel = ws.cell(row=row, column=1, value=shopping_nome.upper())
        cel.font      = Font(size=13, bold=True)
        cel.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = ALTURA_BASE
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = fill_shop
        row += 1

        for loja_id, loja_nome in ordem_lojas[shopping_id]:
            for p in estrutura[shopping_id][loja_id]:
                h = _escrever_linha_pedido(ws, row, loja_nome, p, fonte_padrao, ALTURA_BASE)
                linhas_gastas = max(1, round(h / ALTURA_BASE))
                linhas += linhas_gastas
                row += 1

        for _ in range(ESPACO_SHOP):
            ws.row_dimensions[row].height = ALTURA_BASE
            row += 1
        linhas += ESPACO_SHOP

        return row, linhas  # retorna row E linhas reais

    def _preencher_fim_pagina(row, linhas_usadas):
        faltam = LINHAS_PAGINA - linhas_usadas
        for _ in range(max(0, faltam)):
            ws.row_dimensions[row].height = ALTURA_BASE
            row += 1
        return row

    # ── PRIMEIRA PÁGINA ──
    row_ini_pagina = 1
    row = _escrever_cabecalho_pagina(1)
    linhas_usadas = LINHAS_HEADER

    for shopping_id, shopping_nome in ordem_shoppings:
        tam = _tam_shopping(shopping_id)

        # cabe na página atual?
        if linhas_usadas + tam > LINHAS_UTEIS:
            if linhas_usadas > LINHAS_HEADER:
            # preenche até o fim e quebra
                row = _preencher_fim_pagina(row, linhas_usadas)
                paginas_bordas.append((row_ini_pagina, row - 1))
                ws.row_breaks.append(Break(id=row - 1))

                row_ini_pagina = row
                row = _escrever_cabecalho_pagina(row)
                linhas_usadas = LINHAS_HEADER

        row, tam_real = _escrever_shopping(row, shopping_id, shopping_nome)
        linhas_usadas += tam_real

    # preenche última página
    row = _preencher_fim_pagina(row, linhas_usadas)
    paginas_bordas.append((row_ini_pagina, row - 1))

    for ini, fim in paginas_bordas:
        _aplicar_bordas_pagina(ws, ini, fim, linhas_sem_borda)

    _configurar_impressao(ws, row - 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(buf, as_attachment=True,
                     download_name=_nome_arquivo(viagem),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── SANTA CATARINA ────────────────────────────────────────────────────────────

def _exportar_santa_catarina(viagem, pedidos):
    ALTURA_BASE   = 18
    LINHAS_PAGINA = 42

    estrutura, ordem_shoppings, ordem_lojas = _organizar_estrutura(pedidos)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tarefas"
    _configurar_colunas(ws)

    fonte_padrao = Font(size=13)
    fonte_aviso  = Font(size=11, bold=True, italic=True)
    fill_shop    = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    fill_aviso   = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    linhas_sem_borda = set()

    # header = 2 branco + 1 título + 1 espaço = 4 linhas
    LINHAS_HEADER = 4
    LINHAS_UTIL   = LINHAS_PAGINA - LINHAS_HEADER
    LINHAS_SHOP_1 = LINHAS_UTIL // 2
    LINHAS_SHOP_2 = LINHAS_UTIL - LINHAS_SHOP_1  # pega o resto se ímpar

    def _escrever_bloco_shopping(shopping_id, shopping_nome, row_inicio, linhas_disponiveis):
        r = row_inicio

        cel = ws.cell(row=r, column=1, value=shopping_nome.upper())
        cel.font      = Font(size=13, bold=True)
        cel.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = ALTURA_BASE
        for col in range(1, 5):
            ws.cell(row=r, column=col).fill = fill_shop
        r += 1
        linhas_usadas = 1

        LIMITE = linhas_disponiveis - 1 # reserva 1 pra aviso
        pedidos_nao_exibidos = 0

        for loja_id, loja_nome in ordem_lojas.get(shopping_id, []):
            for p in estrutura[shopping_id][loja_id]:

                if pedidos_nao_exibidos > 0:
                    pedidos_nao_exibidos += 1
                    continue

                h = _escrever_linha_pedido(ws, r, loja_nome, p, fonte_padrao, ALTURA_BASE)
                linhas_gastas = max(1, round(h / ALTURA_BASE))

                if linhas_usadas + linhas_gastas > LIMITE:
                    print(f"  REJEITOU | linhas_usadas={linhas_usadas} linhas_gastas={linhas_gastas} LIMITE={LIMITE}")
                    for col in range(1, 5):
                        ws.cell(row=r, column=col).value     = None
                        ws.cell(row=r, column=col).font      = Font()
                        ws.cell(row=r, column=col).alignment = Alignment()
                    ws.row_dimensions[r].height = ALTURA_BASE
                    pedidos_nao_exibidos += 1
                    continue

                linhas_usadas += linhas_gastas
                r += 1

        if pedidos_nao_exibidos > 0:
            for col in range(1, 5):
                ws.cell(row=r, column=col).value = None
                ws.cell(row=r, column=col).fill  = fill_aviso
            cel_aviso = ws.cell(row=r, column=1,
                                value=f"⚠ +{pedidos_nao_exibidos} pedido(s) não exibido(s)")
            cel_aviso.font      = fonte_aviso
            cel_aviso.alignment = Alignment(vertical="center")
            ws.row_dimensions[r].height = ALTURA_BASE
            r += 1
            linhas_usadas += 1

        while linhas_usadas < linhas_disponiveis:
            ws.row_dimensions[r].height = ALTURA_BASE
            r += 1
            linhas_usadas += 1

        return r

    def _bloco_vazio(row_inicio, linhas):
        r = row_inicio
        for _ in range(linhas):
            ws.row_dimensions[r].height = ALTURA_BASE
            r += 1
        return r

    paginas_bordas = []

    # ── PÁGINA 1 ──────────────────────────────────────────────
    row_ini_pag1 = 1

    for r_b in (1, 2):
        ws.row_dimensions[r_b].height = ALTURA_BASE
        linhas_sem_borda.add(r_b)

    row = 3
    row = _escrever_header(ws, row, viagem, ALTURA_BASE)  # row vira 4

    ws.row_dimensions[row].height = ALTURA_BASE  # 1 espaço após header
    row += 1

    # shop 0
    if len(ordem_shoppings) > 0:
        sid, snome = ordem_shoppings[0]
        row = _escrever_bloco_shopping(sid, snome, row, LINHAS_SHOP_1)
    else:
        row = _bloco_vazio(row, LINHAS_SHOP_1)

    # shop 1
    if len(ordem_shoppings) > 1:
        sid, snome = ordem_shoppings[1]
        row = _escrever_bloco_shopping(sid, snome, row, LINHAS_SHOP_2)
    else:
        row = _bloco_vazio(row, LINHAS_SHOP_2)

    paginas_bordas.append((row_ini_pag1, row - 1))

    # ── PÁGINA 2 ──────────────────────────────────────────────
    ws.row_breaks.append(Break(id=row - 1))
    row_ini_pag2 = row

    for _ in range(2):
        ws.row_dimensions[row].height = ALTURA_BASE
        linhas_sem_borda.add(row)
        row += 1

    row = _escrever_header(ws, row, viagem, ALTURA_BASE)

    ws.row_dimensions[row].height = ALTURA_BASE  # 1 espaço após header
    row += 1

    # shop 2
    if len(ordem_shoppings) > 2:
        sid, snome = ordem_shoppings[2]
        row = _escrever_bloco_shopping(sid, snome, row, LINHAS_SHOP_1)
    else:
        row = _bloco_vazio(row, LINHAS_SHOP_1)

    # shop 3
    if len(ordem_shoppings) > 3:
        sid, snome = ordem_shoppings[3]
        row = _escrever_bloco_shopping(sid, snome, row, LINHAS_SHOP_2)
    else:
        row = _bloco_vazio(row, LINHAS_SHOP_2)

    paginas_bordas.append((row_ini_pag2, row - 1))

    for ini, fim in paginas_bordas:
        _aplicar_bordas_pagina(ws, ini, fim, linhas_sem_borda)

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered   = False
    ws.page_setup.fitToHeight  = False
    ws.page_setup.fitToWidth   = False
    ws.page_setup.fitToPage    = False
    ws.page_setup.scale        = 100
    ws.page_setup.orientation  = ws.ORIENTATION_PORTRAIT
    ws.print_area = f"A1:D{row - 1}"
    ws.page_margins = PageMargins(
        left=0.12, right=0.12, top=0.12, bottom=0.12,
        header=0, footer=0
    )
    ws.oddHeader.left.text   = ""
    ws.oddHeader.center.text = ""
    ws.oddHeader.right.text  = ""
    ws.oddFooter.left.text   = ""
    ws.oddFooter.center.text = ""
    ws.oddFooter.right.text  = ""

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(buf, as_attachment=True,
                     download_name=_nome_arquivo(viagem),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


#exportar ordem
@viagens.route("/viagens/<int:id>/exportar_ordem")
@login_required
def exportar_ordem(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
    SELECT local, data_viagem
    FROM viagens
    WHERE id = %s
""", (id,))
    viagem = cursor.fetchone()

    cursor.execute("""
        SELECT 
            clientes.nome,
            clientes.endereco,
            clientes.telefone,
            vc.ordem
        FROM viagem_clientes vc
        JOIN clientes ON vc.cliente_id = clientes.id
        WHERE vc.viagem_id = %s
        ORDER BY vc.ordem
    """, (id,))

    clientes = cursor.fetchall()


    cursor.close()
    conn.close()

    def limpar_texto(texto):
        return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')

    local = viagem["local"] if viagem and viagem["local"] else "viagem"
    data = viagem["data_viagem"] if viagem and viagem["data_viagem"] else None
    local_str = limpar_texto(local).replace(" ", "_").lower()
    data_str = data.strftime("%d-%m-%Y") if data else "sem_data"

    nome_arquivo = f"OrdemClientes_{local_str}_{data_str}.xlsx"
    altura_padrao = 30

    wb = Workbook()
    ws = wb.active
    ws.title = "Ordem de Clientes"

    # largura das colunas
    ws.column_dimensions['A'].width = 10   # ordem
    ws.column_dimensions['B'].width = 35   # nome
    ws.column_dimensions['C'].width = 20   # endereço
    ws.column_dimensions['D'].width = 20  # telefone
    ws.column_dimensions['E'].width = 20 #horario

    ws.row_dimensions[1].height = altura_padrao
    ws.row_dimensions[2].height = altura_padrao

    row = 3  # ALTERADO - começa na linha 3


    # título
    titulo = ws.cell(row=row, column=1, value="ORDEM DE CLIENTES")
    titulo.font = Font(bold=True, size=16)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = altura_padrao

    row += 1  # vai direto pro cabeçalho

    headers = ["Ordem", "Nome", "Endereço", "Telefone", "Horário"]
    fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    for col, texto in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=texto)
        cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill

    ws.row_dimensions[row].height = altura_padrao

    max_end_len = 0

    row += 1

    # conteúdo
    for i, c in enumerate(clientes, start=1):

        nome = (c["nome"] or "-").upper()
        endereco = c["endereco"]
        telefone = c["telefone"] or "-"
        max_end_len = max(max_end_len, len(endereco))

        # ordem (esquerda)
        cell_ordem = ws.cell(row=row, column=1, value=f"{i}.")
        cell_ordem.font = Font(size=12)
        cell_ordem.alignment = Alignment(horizontal="center", vertical="center")

        # nome (meio)
        cell_nome = ws.cell(row=row, column=2, value=nome)
        cell_nome.font = Font(size=12)
        cell_nome.alignment = Alignment(wrap_text=True, vertical="center")

        # endereço (direita)
        if not endereco or endereco.strip() == "":
            cell_end = ws.cell(row=row, column=3, value="-")
            cell_end.alignment = Alignment(horizontal="center", vertical="center")
            cell_end.font = Font(bold=True)
        else:
            cell_end = ws.cell(row=row, column=3, value=endereco)
            cell_end.alignment = Alignment(wrap_text=True, vertical="center")
            cell_end.font = Font(size=13)

        cell_tel = ws.cell(row=row, column=4, value=telefone)
        cell_tel.font = Font(size=12, bold=True)
        cell_tel.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=5, value="")

        # altura maior pra leitura
        ws.row_dimensions[row].height = altura_padrao

        row += 1
        

    largura_endereco = max(max_end_len * 0.9, 20)
    ws.column_dimensions['C'].width =  min(largura_endereco, 60)
    ws.row_dimensions[row].height = altura_padrao
    ws.print_area = f"A1:E{row}"
    borda = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
    )

    linhas_sem_borda = {1, 2}

    for r in range(1, row):
        if r in linhas_sem_borda:
            continue

        for c in range(1, 6):  # até coluna E
            ws.cell(row=r, column=c).border = borda
    ws.page_margins = PageMargins(
    left=0.10,
    right=0.10,
    top=0.15,
    bottom=0.15
    )
    if len(clientes) <= 20:
        ws.page_setup.fitToHeight = 1
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToPage = True
    else:
        ws.page_setup.fitToPage = False
        ws.page_setup.scale = 100
    file = io.BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(
        file,
        as_attachment=True,
        download_name = nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )